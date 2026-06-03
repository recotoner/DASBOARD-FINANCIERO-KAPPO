from __future__ import annotations

import re
import unicodedata
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd


REQUIRED_COLUMNS = [
    "periodo",
    "grupo",
    "cuenta",
    "monto",
    "origen",
    "nivel",
    "orden",
    "fuente",
]

AUDIT_COLUMNS = ["fila_origen", "monto_origen", "signo_normalizado"]

MONTHS = {
    "ene": 1,
    "enero": 1,
    "feb": 2,
    "febrero": 2,
    "mar": 3,
    "marzo": 3,
    "abr": 4,
    "abril": 4,
    "may": 5,
    "mayo": 5,
    "jun": 6,
    "junio": 6,
    "jul": 7,
    "julio": 7,
    "ago": 8,
    "agosto": 8,
    "sep": 9,
    "sept": 9,
    "septiembre": 9,
    "oct": 10,
    "octubre": 10,
    "nov": 11,
    "noviembre": 11,
    "dic": 12,
    "diciembre": 12,
}

AGGREGATE_HEADERS = {"total", "ano anterior", "ano actual"}
COMPARATIVE_MONTH_RE = re.compile(r"^([a-z]+)\s*-\s*(19\d{2}|20\d{2})$")
DETAIL_CODE_RE = re.compile(r"^\s*\d+(?:\.\d+)+\s*-")


def load_kame_eerr(path_or_file: str | Path | Any) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Load a raw Kame EERR Excel file into the internal Base_normalizada format."""
    wb = openpyxl.load_workbook(path_or_file, data_only=True, read_only=True)
    sheet_name = _select_eerr_sheet(wb)
    ws = wb[sheet_name]
    source_name = _source_name(path_or_file)

    header_row = _detect_header_row(ws)
    year = _detect_year(ws)
    month_columns = _detect_month_columns(ws, header_row, year)

    records: list[dict[str, Any]] = []
    omitted_rows: list[dict[str, Any]] = []
    current_section = ""
    classified_counts = {"detalle": 0, "total": 0, "resultado": 0, "seccion": 0, "omitida": 0}

    for row_idx in range(header_row + 1, ws.max_row + 1):
        row = [_cell_value(ws, row_idx, col_idx) for col_idx in range(1, ws.max_column + 1)]
        if not _has_content(row):
            continue

        row_type = _classify_row(row, month_columns)
        if row_type == "seccion":
            current_section = _clean_text(row[0])
            classified_counts["seccion"] += 1
            continue

        if row_type == "omitida":
            classified_counts["omitida"] += 1
            omitted_rows.append({"fila": row_idx, "motivo": "sin cuenta/montos utiles"})
            continue

        classified_counts[row_type] += 1
        cuenta = _account_name(row, row_type)
        grupo = _normalize_group(current_section, cuenta, row_type)
        orden = row_idx

        for col_idx, periodo in month_columns:
            raw_amount = _to_number(_cell_value(ws, row_idx, col_idx))
            if raw_amount is None:
                raw_amount = 0.0
            monto = _normalize_amount(raw_amount, grupo, row_type)
            records.append(
                {
                    "periodo": periodo,
                    "grupo": grupo,
                    "cuenta": cuenta,
                    "monto": monto,
                    "origen": "kame_eerr",
                    "nivel": row_type,
                    "orden": orden,
                    "fuente": f"{source_name}::{sheet_name}",
                    "fila_origen": row_idx,
                    "monto_origen": raw_amount,
                    "signo_normalizado": bool(raw_amount != monto),
                }
            )

    base = pd.DataFrame(records, columns=REQUIRED_COLUMNS + AUDIT_COLUMNS)
    diagnostics = _build_diagnostics(
        wb=wb,
        sheet_name=sheet_name,
        ws=ws,
        header_row=header_row,
        year=year,
        month_columns=month_columns,
        classified_counts=classified_counts,
        omitted_rows=omitted_rows,
        base=base,
        source_name=source_name,
    )
    return base, diagnostics


def _source_name(path_or_file: str | Path | Any) -> str:
    name = getattr(path_or_file, "name", None)
    if name:
        return Path(str(name)).name
    return Path(str(path_or_file)).name


def _select_eerr_sheet(wb: openpyxl.Workbook) -> str:
    for ws in wb.worksheets:
        first_values = " ".join(
            str(ws.cell(row=r, column=1).value or "") for r in range(1, min(ws.max_row, 8) + 1)
        )
        if "ESTADO" in _fold(first_values) and "RESULTADO" in _fold(first_values):
            return ws.title
    return wb.sheetnames[0]


def _detect_header_row(ws: openpyxl.worksheet.worksheet.Worksheet) -> int:
    best_row = 1
    best_score = -1
    for row_idx in range(1, min(ws.max_row, 25) + 1):
        score = 0
        for col_idx in range(1, ws.max_column + 1):
            value = _clean_text(ws.cell(row=row_idx, column=col_idx).value)
            if _parse_month_header(value) is not None or _is_aggregate_header(value):
                score += 1
        if score > best_score:
            best_score = score
            best_row = row_idx
    if best_score <= 0:
        raise ValueError("No se pudo detectar la fila de meses del Excel Kame.")
    return best_row


def _detect_year(ws: openpyxl.worksheet.worksheet.Worksheet) -> int:
    for row_idx in range(1, min(ws.max_row, 10) + 1):
        for col_idx in range(1, ws.max_column + 1):
            value = _clean_text(ws.cell(row=row_idx, column=col_idx).value)
            match = re.search(r"(20\d{2}|19\d{2})", value)
            if match:
                return int(match.group(1))
    raise ValueError("No se pudo detectar el ano del periodo de emision.")


def _detect_month_columns(
    ws: openpyxl.worksheet.worksheet.Worksheet, header_row: int, year: int
) -> list[tuple[int, str]]:
    columns: list[tuple[int, str]] = []
    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col_idx).value
        if _is_aggregate_header(value):
            continue
        parsed = _parse_month_header(value)
        if parsed is None:
            continue
        month, parsed_year = parsed
        period_year = parsed_year if parsed_year is not None else year
        columns.append((col_idx, f"{period_year}-{month:02d}"))
    if not columns:
        raise ValueError("No se detectaron columnas mensuales en el Excel Kame.")
    return columns


def _parse_month_header(value: Any) -> tuple[int, int | None] | None:
    folded = _fold(value)
    if not folded:
        return None
    if folded in MONTHS:
        return MONTHS[folded], None
    normalized = re.sub(r"\s+", " ", folded)
    match = COMPARATIVE_MONTH_RE.match(normalized)
    if not match:
        return None
    month = MONTHS.get(match.group(1))
    if not month:
        return None
    return month, int(match.group(2))


def _is_aggregate_header(value: Any) -> bool:
    return _fold(value) in AGGREGATE_HEADERS


def _classify_row(row: list[Any], month_columns: list[tuple[int, str]]) -> str:
    col_a = _clean_text(row[0] if len(row) > 0 else "")
    col_b = _clean_text(row[1] if len(row) > 1 else "")
    has_amounts = any(_to_number(row[col_idx - 1]) is not None for col_idx, _ in month_columns)

    if not col_a and not col_b and not has_amounts:
        return "omitida"
    if col_a and not col_b and not has_amounts:
        return "seccion"
    if _fold(col_a).startswith("total") and has_amounts:
        return "total"
    if DETAIL_CODE_RE.match(col_b) and has_amounts:
        return "detalle"
    if col_b and has_amounts:
        return "resultado"
    return "omitida"


def _account_name(row: list[Any], row_type: str) -> str:
    if row_type == "total":
        return _clean_text(row[0])
    return _clean_text(row[1])


def _normalize_group(section: str, cuenta: str, level: str) -> str:
    if level == "resultado":
        return "Resultado"

    text = _fold(f"{section} {cuenta}")
    cuenta_text = _fold(cuenta)
    if re.match(r"^4\.02\.07(?:\.|$)", _clean_text(cuenta)) or "gastos bancarios" in cuenta_text:
        return "Gastos financieros"
    if "ingreso" in text and "egreso" not in text:
        if "explotacion" in text:
            return "Ingresos de explotacion"
        return "Otros ingresos"
    if "costo" in text:
        return "Costos de explotacion"
    if "gasto" in text and "financiero" in text:
        return "Gastos financieros"
    if "gasto" in text:
        return "Gastos administracion y ventas"
    if "sueldo" in text or "remuneracion" in text or "leyes sociales" in text:
        return "Sueldos y leyes sociales"
    if "egreso" in text or "perdida" in text:
        return "Otros egresos"
    if "depreciacion" in text:
        return "Depreciaciones"
    if "correccion monetaria" in text:
        return "Correccion monetaria"
    if "impuesto" in text:
        return "Impuesto a la renta"
    return _clean_text(section) or "Sin grupo"


def _normalize_amount(raw_amount: float, group: str, level: str) -> float:
    if level == "resultado":
        return float(raw_amount)
    group_folded = _fold(group)
    expense_tokens = [
        "costo",
        "gasto",
        "sueldo",
        "egreso",
        "depreciacion",
        "correccion",
        "impuesto",
    ]
    if any(token in group_folded for token in expense_tokens):
        if raw_amount > 0:
            return -float(raw_amount)
        if raw_amount < 0:
            return abs(float(raw_amount))
        return 0.0
    return float(raw_amount)


def _build_diagnostics(
    *,
    wb: openpyxl.Workbook,
    sheet_name: str,
    ws: openpyxl.worksheet.worksheet.Worksheet,
    header_row: int,
    year: int,
    month_columns: list[tuple[int, str]],
    classified_counts: dict[str, int],
    omitted_rows: list[dict[str, Any]],
    base: pd.DataFrame,
    source_name: str,
) -> dict[str, Any]:
    sign_changes = int(base["signo_normalizado"].sum()) if not base.empty else 0
    filas_base_por_nivel = (
        base["nivel"].value_counts().sort_index().astype(int).to_dict() if not base.empty else {}
    )
    control_cuadratura = _build_control_cuadratura(base)
    resumen_control_cuadratura = _build_resumen_control_cuadratura(control_cuadratura)
    diferencias_a_revisar = [row for row in control_cuadratura if not row["cuadra"]]
    exceptions = _build_standard_exceptions(
        diferencias_a_revisar=diferencias_a_revisar,
        adapter="kame_eerr",
        fuente=source_name,
    )
    detalle_montos_negativos_diferencias = _build_detalle_montos_negativos_diferencias(
        base, diferencias_a_revisar
    )
    negative_expense_rows = []
    if not base.empty:
        mask = (base["monto_origen"] < 0) & (base["nivel"].isin(["detalle", "total"]))
        negative_expense_rows = (
            base.loc[mask, ["fila_origen", "grupo", "cuenta", "periodo", "monto_origen"]]
            .drop_duplicates()
            .to_dict("records")
        )

    return {
        "archivo": source_name,
        "hojas": wb.sheetnames,
        "hoja_usada": sheet_name,
        "dimension": f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}",
        "filas": ws.max_row,
        "columnas": ws.max_column,
        "fila_header_meses": header_row,
        "anio_detectado": year,
        "periodos_detectados": [period for _, period in month_columns],
        "columnas_meses": [
            {"columna_excel": _column_letter(col_idx), "periodo": period}
            for col_idx, period in month_columns
        ],
        "columna_total_ignorada": True,
        "conteo_filas_clasificadas": classified_counts,
        "filas_base_por_nivel": filas_base_por_nivel,
        "filas_base_normalizada": int(len(base)),
        "columnas_base_normalizada": REQUIRED_COLUMNS,
        "columnas_auditoria": AUDIT_COLUMNS,
        "control_cuadratura": control_cuadratura,
        "resumen_control_cuadratura": resumen_control_cuadratura,
        "diferencias_a_revisar": diferencias_a_revisar,
        "exceptions": exceptions,
        "detalle_montos_negativos_diferencias": detalle_montos_negativos_diferencias,
        "cambios_de_signo": sign_changes,
        "montos_origen_negativos": negative_expense_rows,
        "filas_omitidas_muestra": omitted_rows[:20],
    }


def _build_control_cuadratura(base: pd.DataFrame) -> list[dict[str, Any]]:
    if base.empty:
        return []

    levels_by_group = base.groupby("grupo")["nivel"].agg(lambda values: set(values))
    groups_with_detail_and_total = [
        group
        for group, levels in levels_by_group.items()
        if {"detalle", "total"}.issubset(levels)
    ]
    if not groups_with_detail_and_total:
        return []

    eligible = base[
        base["grupo"].isin(groups_with_detail_and_total)
        & base["nivel"].isin(["detalle", "total"])
    ].copy()
    pivot = (
        eligible.groupby(["periodo", "grupo", "nivel"], as_index=False)["monto"]
        .sum()
        .pivot_table(
            index=["periodo", "grupo"],
            columns="nivel",
            values="monto",
            aggfunc="sum",
            fill_value=0.0,
        )
        .reset_index()
    )
    if "detalle" not in pivot.columns:
        pivot["detalle"] = 0.0
    if "total" not in pivot.columns:
        pivot["total"] = 0.0

    pivot = pivot.rename(
        columns={
            "detalle": "suma_detalle",
            "total": "total_kame_normalizado",
        }
    )
    pivot["diferencia"] = pivot["suma_detalle"] - pivot["total_kame_normalizado"]
    pivot["cuadra"] = pivot["diferencia"].abs() <= 1
    pivot = pivot.sort_values(["periodo", "grupo"])
    return pivot[
        ["periodo", "grupo", "suma_detalle", "total_kame_normalizado", "diferencia", "cuadra"]
    ].to_dict("records")


def _build_resumen_control_cuadratura(control_cuadratura: list[dict[str, Any]]) -> dict[str, Any]:
    grupos_periodo_auditados = len(control_cuadratura)
    controles_ok = sum(1 for row in control_cuadratura if row["cuadra"])
    controles_con_diferencia = grupos_periodo_auditados - controles_ok
    mayor_diferencia_abs = (
        max(abs(float(row["diferencia"])) for row in control_cuadratura)
        if control_cuadratura
        else 0.0
    )
    return {
        "grupos_periodo_auditados": grupos_periodo_auditados,
        "controles_ok": controles_ok,
        "controles_con_diferencia": controles_con_diferencia,
        "mayor_diferencia_abs": mayor_diferencia_abs,
        "estado_general": "OK" if controles_con_diferencia == 0 else "REVISAR",
    }


def _build_standard_exceptions(
    *,
    diferencias_a_revisar: list[dict[str, Any]],
    adapter: str,
    fuente: str,
) -> list[dict[str, Any]]:
    exceptions: list[dict[str, Any]] = []
    for row in diferencias_a_revisar:
        periodo = str(row.get("periodo", ""))
        grupo = str(row.get("grupo", ""))
        tipo_excepcion = "group_total_reconciliation_mismatch"
        exception_id = _build_exception_id(
            adapter=adapter,
            fuente=fuente,
            periodo=periodo,
            grupo=grupo,
            tipo_excepcion=tipo_excepcion,
        )
        diferencia = float(row.get("diferencia", 0) or 0)
        exceptions.append(
            {
                "exception_id": exception_id,
                "adapter": adapter,
                "fuente": fuente,
                "periodo": periodo,
                "grupo": grupo,
                "cuenta": None,
                "control_tipo": "group_total_reconciliation",
                "tipo_excepcion": tipo_excepcion,
                "severidad": "pendiente_clasificacion",
                "estado": "pendiente",
                "diferencia": diferencia,
                "mensaje": (
                    f"El grupo '{grupo}' no cuadra en el periodo {periodo}: "
                    f"suma detalle {row.get('suma_detalle')} vs total Kame "
                    f"{row.get('total_kame_normalizado')}."
                ),
                "contexto": {
                    "suma_detalle": row.get("suma_detalle"),
                    "total_kame_normalizado": row.get("total_kame_normalizado"),
                    "cuadra": row.get("cuadra"),
                },
            }
        )
    return exceptions


def _build_exception_id(
    *,
    adapter: str,
    fuente: str,
    periodo: str,
    grupo: str,
    tipo_excepcion: str,
) -> str:
    raw = f"{adapter}__{fuente}__{periodo}__{grupo}__{tipo_excepcion}"
    return re.sub(r"[^a-zA-Z0-9_\\-]+", "_", _fold(raw)).strip("_").lower()


def _build_detalle_montos_negativos_diferencias(
    base: pd.DataFrame, diferencias_a_revisar: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    if base.empty or not diferencias_a_revisar:
        return []

    keys = {(row["periodo"], row["grupo"]) for row in diferencias_a_revisar}
    mask = base.apply(lambda row: (row["periodo"], row["grupo"]) in keys, axis=1)
    detail = base[
        mask
        & (base["nivel"] == "detalle")
        & (base["monto_origen"] < 0)
    ].copy()
    if detail.empty:
        return []
    return detail[
        [
            "periodo",
            "grupo",
            "cuenta",
            "fila_origen",
            "monto_origen",
            "monto",
            "signo_normalizado",
        ]
    ].sort_values(["periodo", "grupo", "fila_origen"]).to_dict("records")


def _cell_value(ws: openpyxl.worksheet.worksheet.Worksheet, row_idx: int, col_idx: int) -> Any:
    return ws.cell(row=row_idx, column=col_idx).value


def _has_content(row: list[Any]) -> bool:
    return any(_clean_text(value) for value in row)


def _to_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _clean_text(value)
    if not text:
        return None
    text = text.replace("$", "").replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _fold(value: Any) -> str:
    text = _clean_text(value).lower()
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def _column_letter(col_idx: int) -> str:
    return openpyxl.utils.get_column_letter(col_idx)

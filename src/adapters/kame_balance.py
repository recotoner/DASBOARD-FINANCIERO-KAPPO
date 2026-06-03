from __future__ import annotations

import re
import unicodedata
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd


REQUIRED_COLUMNS = [
    "periodo",
    "lado",
    "grupo",
    "subgrupo",
    "cuenta",
    "codigo",
    "monto",
    "origen",
    "nivel",
    "orden",
    "fuente",
]

AUDIT_COLUMNS = [
    "fila_origen",
    "columna_origen",
    "monto_origen",
    "signo_normalizado",
]

CODE_RE = re.compile(r"^\s*(\d+(?:\.\d+)*)\s*-\s*(.+)$")
DATE_RE = re.compile(r"(\d{2})-(\d{2})-(20\d{2}|19\d{2})")


def load_kame_balance(
    path_or_file: str | Path | Any, periodo: str | None = None
) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Load a raw Kame classified balance into Base_balance_normalizada."""
    wb = openpyxl.load_workbook(path_or_file, data_only=True, read_only=True)
    sheet_name = _select_balance_sheet(wb)
    ws = wb[sheet_name]
    source_name = _source_name(path_or_file)
    periodo_detectado = periodo or _infer_period_from_source(source_name)

    records: list[dict[str, Any]] = []
    records.extend(
        _read_balance_block(
            ws=ws,
            group_col=1,
            detail_col=2,
            amount_col=3,
            side_hint="activo",
            periodo=periodo_detectado,
            source_name=source_name,
            sheet_name=sheet_name,
        )
    )
    records.extend(
        _read_balance_block(
            ws=ws,
            group_col=5,
            detail_col=6,
            amount_col=7,
            side_hint="pasivo_patrimonio",
            periodo=periodo_detectado,
            source_name=source_name,
            sheet_name=sheet_name,
        )
    )

    base = pd.DataFrame(records, columns=REQUIRED_COLUMNS + AUDIT_COLUMNS)
    diagnostics = _build_diagnostics_balance(
        wb=wb,
        ws=ws,
        sheet_name=sheet_name,
        source_name=source_name,
        periodo=periodo_detectado,
        base=base,
    )
    return base, diagnostics


def _read_balance_block(
    *,
    ws: openpyxl.worksheet.worksheet.Worksheet,
    group_col: int,
    detail_col: int,
    amount_col: int,
    side_hint: str,
    periodo: str,
    source_name: str,
    sheet_name: str,
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    current_group = ""
    current_subgroup = ""
    current_code = ""
    current_lado = "activo" if side_hint == "activo" else "pasivo"

    for row_idx in range(2, ws.max_row + 1):
        group_text = _clean_text(ws.cell(row=row_idx, column=group_col).value)
        detail_text = _clean_text(ws.cell(row=row_idx, column=detail_col).value)
        raw_amount = _to_number(ws.cell(row=row_idx, column=amount_col).value)

        if not group_text and not detail_text and raw_amount is None:
            continue

        if _fold(group_text) == "resultado del ejercicio":
            records.append(
                _record(
                    periodo=periodo,
                    lado="resultado",
                    grupo="Resultado",
                    subgrupo="Resultado del ejercicio",
                    cuenta="Resultado del ejercicio",
                    codigo="",
                    monto=raw_amount or 0.0,
                    nivel="resultado",
                    row_idx=row_idx,
                    amount_col=amount_col,
                    source_name=source_name,
                    sheet_name=sheet_name,
                )
            )
            current_lado = "resultado"
            current_group = "Resultado"
            current_subgroup = "Resultado del ejercicio"
            current_code = ""
            continue

        parsed = _parse_code_line(group_text)
        if parsed:
            code, label = parsed
            lado, group = _classify_side_group(code, label, side_hint)
            current_lado = lado
            current_group = group
            current_subgroup = label
            current_code = code
            if raw_amount is not None:
                records.append(
                    _record(
                        periodo=periodo,
                        lado=lado,
                        grupo=group,
                        subgrupo=label,
                        cuenta=label,
                        codigo=code,
                        monto=raw_amount,
                        nivel="total",
                        row_idx=row_idx,
                        amount_col=amount_col,
                        source_name=source_name,
                        sheet_name=sheet_name,
                    )
                )
            continue

        if detail_text and raw_amount is not None:
            records.append(
                _record(
                    periodo=periodo,
                    lado=current_lado,
                    grupo=current_group,
                    subgrupo=current_subgroup,
                    cuenta=detail_text,
                    codigo=current_code,
                    monto=raw_amount,
                    nivel="detalle",
                    row_idx=row_idx,
                    amount_col=amount_col,
                    source_name=source_name,
                    sheet_name=sheet_name,
                )
            )

    return records


def _build_diagnostics_balance(
    *,
    wb: openpyxl.Workbook,
    ws: openpyxl.worksheet.worksheet.Worksheet,
    sheet_name: str,
    source_name: str,
    periodo: str,
    base: pd.DataFrame,
) -> dict[str, Any]:
    control = _build_control_balance(base)
    filas_por_lado = (
        base["lado"].value_counts().sort_index().astype(int).to_dict()
        if not base.empty
        else {}
    )
    filas_por_nivel = (
        base["nivel"].value_counts().sort_index().astype(int).to_dict()
        if not base.empty
        else {}
    )
    return {
        "archivo": source_name,
        "hojas": wb.sheetnames,
        "hoja_usada": sheet_name,
        "dimension": f"A1:{_column_letter(ws.max_column)}{ws.max_row}",
        "periodo": periodo,
        "filas": ws.max_row,
        "columnas": ws.max_column,
        "filas_base_balance": int(len(base)),
        "filas_por_lado": filas_por_lado,
        "filas_por_nivel": filas_por_nivel,
        "columnas_base_balance": REQUIRED_COLUMNS,
        "columnas_auditoria": AUDIT_COLUMNS,
        "control_balance": control,
    }


def _build_control_balance(base: pd.DataFrame) -> dict[str, Any]:
    total_activos = _amount_by_code(base, "1")
    pasivo_corriente = _amount_by_code(base, "2.01")
    patrimonio = _amount_by_code(base, "2.03")
    resultado = _amount_by_level_group(base, "resultado", "Resultado")
    total_derecha = pasivo_corriente + patrimonio + resultado
    diferencia = total_activos - total_derecha
    return {
        "total_activos": total_activos,
        "pasivo_corriente": pasivo_corriente,
        "patrimonio": patrimonio,
        "resultado_ejercicio": resultado,
        "pasivo_patrimonio_resultado": total_derecha,
        "diferencia_balance": diferencia,
        "cuadra_balance": abs(diferencia) <= 1,
    }


def _amount_by_code(base: pd.DataFrame, code: str) -> float:
    if base.empty:
        return 0.0
    rows = base[(base["nivel"] == "total") & (base["codigo"].astype(str) == code)]
    if rows.empty:
        return 0.0
    return float(rows.iloc[0]["monto"])


def _amount_by_level_group(base: pd.DataFrame, level: str, group: str) -> float:
    if base.empty:
        return 0.0
    rows = base[(base["nivel"] == level) & (base["grupo"] == group)]
    if rows.empty:
        return 0.0
    return float(rows["monto"].sum())


def _classify_side_group(code: str, label: str, side_hint: str) -> tuple[str, str]:
    folded = _fold(label)
    if code.startswith("1.01"):
        return "activo", "Activos Circulantes"
    if code.startswith("1.02"):
        return "activo", "Activos Fijos"
    if code.startswith("1.03"):
        return "activo", "Otros Activos"
    if code.startswith("1"):
        return "activo", "Activos"
    if code.startswith("2.01"):
        return "pasivo", "Pasivos Circulantes"
    if code.startswith("2.02"):
        return "pasivo", "Pasivos No Corrientes"
    if code.startswith("2.03") or "patrimonio" in folded:
        return "patrimonio", "Patrimonio"
    if code.startswith("2"):
        return "pasivo", "Pasivos"
    return ("activo", "Activos") if side_hint == "activo" else ("pasivo", "Pasivos")


def _record(
    *,
    periodo: str,
    lado: str,
    grupo: str,
    subgrupo: str,
    cuenta: str,
    codigo: str,
    monto: float,
    nivel: str,
    row_idx: int,
    amount_col: int,
    source_name: str,
    sheet_name: str,
) -> dict[str, Any]:
    return {
        "periodo": periodo,
        "lado": lado,
        "grupo": grupo,
        "subgrupo": subgrupo,
        "cuenta": cuenta,
        "codigo": codigo,
        "monto": float(monto),
        "origen": "kame_balance",
        "nivel": nivel,
        "orden": row_idx,
        "fuente": f"{source_name}::{sheet_name}",
        "fila_origen": row_idx,
        "columna_origen": _column_letter(amount_col),
        "monto_origen": float(monto),
        "signo_normalizado": False,
    }


def _select_balance_sheet(wb: openpyxl.Workbook) -> str:
    for ws in wb.worksheets:
        first_values = " ".join(
            str(ws.cell(row=r, column=c).value or "")
            for r in range(1, min(ws.max_row, 5) + 1)
            for c in range(1, min(ws.max_column, 7) + 1)
        )
        folded = _fold(first_values)
        if "cuenta activo" in folded and "cuenta pasivo" in folded:
            return ws.title
    return wb.sheetnames[0]


def _source_name(path_or_file: str | Path | Any) -> str:
    name = getattr(path_or_file, "name", None)
    if name:
        return Path(str(name)).name
    return Path(str(path_or_file)).name


def _infer_period_from_source(source_name: str) -> str:
    match = DATE_RE.search(source_name)
    if not match:
        return "sin_periodo"
    _, month, year = match.groups()
    return f"{year}-{month}"


def _parse_code_line(value: Any) -> tuple[str, str] | None:
    match = CODE_RE.match(_clean_text(value))
    if not match:
        return None
    return match.group(1), _clean_text(match.group(2))


def _to_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _clean_text(value)
    if not text:
        return None
    text = text.replace("$", "").replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _fold(value: Any) -> str:
    text = _clean_text(value).lower()
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def _column_letter(col_idx: int) -> str:
    return openpyxl.utils.get_column_letter(col_idx)
